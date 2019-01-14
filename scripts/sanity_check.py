"""
SANITY CHECK script
by M. Amadio (mattia.amadio@gmail.com)
Python coding by T. Iwanaga (iwanaga.takuya@anu.edu.au)

This script computes zonal statistics at ADM0 and ADM1 level on provided global
hazard layers and exports as multisheet xls.

List of Hazards checked
V	 means data are compatible
-	 means data should be reformatted, e.g. tsunami point lay out of boundaries

V	EQ Earthquale
-	TS Tsunami
-	VO Volcano
V	LS Landslide
V	FL Flood
V	SS Storm Surge
V	SW Strong winds
-	WS Water scarcity
V	ET Extreme temperatures
V	WF Wildfires

USAGE
Requires python 3.6 and the following libraries (conda-forge):
- geopandas
- rasterio
- rasterstats
Make sure all layers are in crs wgs84 and rasters are single band.
The rasterstats analysis is based on the raster resolution.

"""
import os
import pathlib
from glob import glob
from datetime import datetime
import itertools as itools

import numpy as np
import pandas as pd
import geopandas as gp
from openpyxl import load_workbook

import rasterio
from rasterstats import zonal_stats

try:
    # Optional package for progress bar
    from tqdm import tqdm
except ImportError:
    pass


def extract_stats(shp_files, rst_files, adms_to_check,
                  stats_of_interest, output_fn):
    """Extract statistics from rasters using shapefiles.

    Parameters
    ==========
    * shp_file : list[str], paths to shapefiles
    * rst_files : list[str], paths to raster files (in .tif format)
    * adms_to_check : list[str], ADM levels to check (["ADM0", "ADM1", ...])
    * stats_of_interest : str, listing statistical functions of interest
                          `zonal_stats()` compatible

    Returns
    ==========
    * None

    Outputs
    ==========
    * Excel File : `sanity_check_[datetime of run].xlsx`
    """
    # combination of all the three items. * expands the list of inputs
    all_combinations = itools.product(*[shp_files, rst_files, adms_to_check])
    # creates empty dictionaries
    shp_cache = {}
    rst_cache = {}
    for shp, rst, adm in all_combinations:
        # Gets raster file name
        sheet_name = os.path.basename(rst)
        shp_name = os.path.basename(shp)

        print("Processing", shp_name, sheet_name, adm)

        # Get number from ADM level string
        adm_code_val = int(adm.strip('ADM'))
        adm_code = "{}_CODE".format(adm)
        adm_name = "{}_NAME".format(adm)

        # Keep shp data in memory until done with it
        if shp in shp_cache:
            shp_data = shp_cache[shp]
        else:
            shp_cache = {}  # clear shapefile cache
            shp_data = gp.read_file(shp).set_index(adm_code)
            
            # Both vector and raster files have to have matching CRS
            with rasterio.open(rst) as src:
                shp_crs = shp_data.crs['init']
                rst_crs = src.crs.to_string().split("=")[1]

                crs_no_match = shp_crs != rst_crs
                rst_not_wgs84 = "4326" not in rst_crs
                shp_not_wgs84 = "4326" not in shp_crs
                if crs_no_match or rst_not_wgs84 or shp_not_wgs84:
                    # CRS do not match so write out error to
                    # xlsx file and then continue
                    with pd.ExcelWriter(output_fn, engine='openpyxl')\
                            as writer:
                        df = pd.DataFrame()
                        df.to_excel(writer, sheet_name, startrow=2)
                        del df

                        cmt = "Could not process {} with {}, incorrect CRS."\
                              .format(sheet_name, shp_name)
                        issue = "Issues:"
                        if crs_no_match:
                            issue += " Shapefile and raster have differing CRS."
                        if shp_not_wgs84:
                            issue += " Shapefile is not in WGS84."
                        if rst_not_wgs84:
                            issue += " Raster is not in WGS84."
                        worksheet = writer.sheets[sheet_name]
                        worksheet.cell(row=1, column=1).value = cmt
                        worksheet.cell(row=2, column=1).value = issue
                    continue
                # End if

# sort adm zones by ascending area size
            shp_data['area'] = shp_data.geometry.area
            shp_data = shp_data.sort_values('area')
            shp_cache[shp] = shp_data

# keep raster in memory until we are done with it
        if rst in rst_cache:
            rst_data = rst_cache[rst]
            nodata = rst_cache['nodata']
            transform = rst_cache['transform']
        else:
            rst_cache = {}  # clear raster cache
            with rasterio.open(rst) as src:
                nodata = src.nodatavals
                transform = src.transform
                rst_data = src.read(1)
                rst_data = np.ma.masked_array(rst_data, mask=(rst_data == nodata))
                rst_cache[rst] = {
                    'data': rst_data,
                    'transform': src.transform,
                    'nodata': nodata
                }
            # End with
        # End if

        # shows progress
        try:
            _loop = tqdm(shp_data.index.unique())
        except NameError:
            _loop = shp_data.index.unique()
        for _admcode in _loop:
            # Using shp_data.loc[adm_code, :] results in recursion error
            # when converting to JSON, so have to explicitly subset DF
            sel = shp_data.loc[shp_data.index == _admcode, :]
            
            # Combine (dissolve) areas with the same ADM code together
            if len(sel.index) > 1:
                sel = sel.dissolve(adm_code)

            curr_adm_name = sel.at[_admcode, adm_name]

            try:
                # Set msg for progress bar if available
                _loop.set_description(curr_adm_name)
            except AttributeError:
                pass

            #creates dictionary of statistics for each individual adm unit
            stats = zonal_stats(sel, rst_data,
                                affine=transform,
                                stats=stats_of_interest,
                                nodata=nodata)
            stats = stats[0]
            stats_for_raster = {adm_name: curr_adm_name}

            # checks for smaller administrative units than ADM0 and list them
            if adm_code_val > 0:
                stats_for_raster['ADM0_CODE'] = shp_data.at[_admcode,
                                                            'ADM0_CODE']

            stats_for_raster.update(stats)
            stats_for_raster.update({'Comment/Error': ""})

            # Add new row to dataframe, or create dataframe
            # if it does not exist
            try:
                df.loc[_admcode] = stats_for_raster
            except NameError:
                df = pd.DataFrame(columns=[col for col in stats_for_raster],
                                  index=shp_data.index.unique())
                df.index.name = adm_code

                df.loc[_admcode] = stats_for_raster
            # End with
        # End adm loop

        # Write out dataframe to excel file
        try:
            book = load_workbook(output_fn)
        except FileNotFoundError:
            book = None

        # Write out data to excel file
        with pd.ExcelWriter(output_fn, engine='openpyxl') as writer:
            if book:
                writer.book = book

            df.to_excel(writer, sheet_name, startrow=2)
            comment = "# Extracted from {} using {} on {}"\
                      .format(sheet_name,
                              shp_name,
                              run_date)
            worksheet = writer.sheets[sheet_name]
            worksheet.cell(row=1, column=1).value = comment
    # End combination loop
# End main()


if __name__ == '__main__':
    run_date = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    output_fn = "sanity_check_{}.xlsx".format(run_date)

    # openpyxl has trouble with relative paths
    # convert to absolute path to avoid this issue
    abs_output = pathlib.Path(output_fn)
    abs_output = abs_output.absolute().resolve()
    print("Outputting data to", abs_output)

    # Shapefile for stats zones
    shp_file_pth = "../data/g2015_2014_0_upd270117"

    # More than one raster is stored in hazards. It gets the list of tif files.
    hazpath = "../data/hazards"

    # Get geotiff and shapefiles
    haz_rasters = glob(hazpath + "/*.tif")
    shp_files = glob(shp_file_pth + "/*.shp")

    assert len(shp_files) > 0, "No shapefiles found!"
    assert len(haz_rasters) > 0, "No rasters found!"

    # select statistics of interest and administrative units
    stats_of_interest = "max min mean range count"
    adms_to_check = ["ADM0"]

    extract_stats(shp_files, haz_rasters, adms_to_check,
                  stats_of_interest, abs_output)
