"""

HAZARD STATISTICS script
by M. Amadio (mattia.amadio@gmail.com)
Python coding by T. Iwanaga (iwanaga.takuya@anu.edu.au)

Script to run zonal statistics on batch rasters.

USAGE
- Have 1 shapefile to use for zonal extraction and 1 or more raster in a data folder.
- THE SHAPEFILE AND THE RASTERS MUST ALL BE THE SAME PROJECTED (meters) CRS.
- Run qa_runstats.py and answer the prompt questions. 
  The tool will select autmatically the one shp and all rasters found in the folder. 
  The ouput xlsx is generated in the data folder.
- Use runstats.py for direct commandline execution.
"""

import rasterio
from rasterio.mask import mask
from rasterio.crs import CRS
import shapely
from fuzzywuzzy import fuzz

import numpy as np
import pandas as pd
import geopandas as gpd

from openpyxl import load_workbook

try:
    from tqdm import tqdm
except ImportError:
    pass

import itertools as itools
from glob import glob
import os


def filter_geometries(shp, rst):
    """Filter geometries to just those that overlay the given raster.

    Parameters
    ==========
    * shp : GeoPandas DataFrame, shapefile data
    * rst : rasterio raster object, raster data

    Returns
    =======
    * GeoPandas DataFrame of geometries inside raster bounds
    """
    # Filter shapefile to geometries inside raster bounds
    minx, miny, maxx, maxy = rst.bounds
    bound_box = shapely.geometry.Polygon([(minx, miny), (minx, maxy), (maxx, maxy), (maxx, miny)])
    relevant_geoms = shp[shp.geometry.within(bound_box)]

    return relevant_geoms
# End filter_geometries()


def mp_calc_stats(in_data, preprocess=None):
    """Calculate given statistics and populate result dict.

    Compatible with multiprocessing methods.

    Parameters
    ==========
    * in_data : tuple, (Pandas row, str, rasterio object, list, dict)
    * preprocess : tuple[str, float], kind of preprocess with associated threshold value.
                   ('SD', 2) : Filter values outside of 2 standard deviations
                   ('PC', 80) : Filter values above 80th percentile
    """
    (row, field, ds, stats, result_set) = in_data

    clip, out_transform = mask(ds, [row.geometry], crop=True)
    clip = np.extract(clip != ds.nodata, clip)

    run_range = 'range' in stats
    if run_range:
        stats = [i for i in stats[:] if i != 'range']

    # Early exit if everything is NaN
    if np.isnan(clip).all():
        return

    # Apply preprocess if necessary
    if preprocess:
        kind, value = preprocess
        if kind == 'SD':
            sd = clip.std()
            avg = clip.mean()
            min_thres = (clip >= (avg - (sd * value)))
            max_thres = (clip <= (avg + (sd * value)))
            clip = clip[min_thres & max_thres]
        elif kind == 'PC':
            p_threshold = np.percentile(clip, value)
            clip = clip[clip <= p_threshold]
        # End if
    # End if
    
    del ds
    res = {func: float(getattr(clip, func)()) for func in stats}

    if run_range:
        res['range'] = float(clip.max() - clip.min())

    if any(res.values()):
        idx = getattr(row, field)
        result_set[idx] = res

# End mp_calc_stats()


def matching_crs(rst, shp_data):
    """Check that raster and shpfile have matching CRS.

    Parameters
    ==========
    * rst : file pointer, pointer to raster file
    * shp_data : geopandas, shape data

    Returns
    =======
    * Tuple[bool, str] : (True/False indicating CRS match, 
                          and str for failure)
    """
    # Both vector and raster files have to have matching CRS
    shp_str = shp_data.crs['init']

    rst_crs = rst.crs
    if rst_crs.is_epsg_code:
        rst_str = rst_crs.to_string()
        if "=" in rst_str:
            rst_str = rst_str.split("=")[1]

        crs_no_match = shp_str.lower() != rst_str.lower()
    else:
        # ...could not identify EPSG code, relying on fuzzy matching of WKT...
        rst_str = rst_crs.to_wkt()
        shp_str = CRS.to_wkt(CRS.from_string(shp_str))

        shp_str = shp_str.replace('[', '').replace(']', '')
        rst_str = rst_str.replace('[', '').replace(']', '')

        # WKT string has to be at least 70% similar...
        crs_no_match = fuzz.token_sort_ratio(shp_str, rst_str) < 70
    # End if

    if crs_no_match:
        # CRS does not match
        issue = "Issues:"
        issue += " Shapefile and raster have differing CRS."
        issue += " shpfile: {}, raster: {}".format(shp_str, rst_str)
        
        return (False, issue)
    # End if

    return (True, '')
# End matching_crs()


def write_to_excel(output_fn, results):
    """Write data to a worksheet.

    Parameters
    ==========
    * output_fn : file pointer, pointer to raster file
    * results : dict[tuple], of statistic in 
                (geopandas df, sheet_name (str), comment (str))
    """
    try:
        book = load_workbook(output_fn)
        mode = 'a'
    except FileNotFoundError:
        book = None
        mode = 'w'

    # Write out data to excel file
    with pd.ExcelWriter(output_fn, engine='openpyxl', mode=mode) as writer:
        if book is not None:
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}

        # data, sheet_name, comment, msg=''
        for data, sheet, comment in results.values():
            # write out data first
            if data is not None:
                data.to_excel(writer, sheet, startrow=2, index=False)
            else:
                df = pd.DataFrame()
                df.to_excel(writer, sheet, startrow=2)

            if comment != '':
                # then add comment (sheet has to exist first, hence the write out above)
                worksheet = writer.sheets[sheet]
                worksheet.cell(row=1, column=1).value = comment

    # End with
# End write_to_excel()


def extract_stats(shp_files, rasters, field, stats, preprocess=None):
    """Extract statistics from a raster using a shapefile.

    Writes results to the specified output file.

    Parameters
    ==========
    * shp_files : List[str], of file paths to shapefiles.
    * rasters : List[str], of file paths to rasters.
    * field : str, name of column to use as primary id.
    * stats : List[str], of statistics to calculate. 
              Numpy compatible method names only.
    """
    all_combinations = itools.product(*[shp_files, rasters])

    results = {}
    shp_cache = {}
    for shp_fn, rst_fn in all_combinations:
        if shp_fn in shp_cache:
            shp = shp_cache[shp_fn].copy()
        else:
            shp_cache = {}  # clear cache
            shp = gpd.read_file(shp_fn)
            shp_cache[shp_fn] = shp.copy()
        # End if
        
        ds = rasterio.open(rst_fn)

        sheet_name = os.path.basename(rst_fn)
        shp_name = os.path.basename(shp_fn)

        crs_matches, issue = matching_crs(ds, shp)
        if not crs_matches:
            cmt = "Could not process {} with {}.\n{}"\
                    .format(sheet_name, shp_name, issue)
            results[(shp_fn, rst_fn)] = None, sheet_name, cmt
            continue
        # End if

        # Filter shapefile to geometries inside raster bounds
        relevant_geoms = filter_geometries(shp, ds)

        try:
            # Set msg for progress bar if available
            _loop = tqdm(relevant_geoms.itertuples())
        except NameError:
            _loop = relevant_geoms.itertuples()

        result_set = {}
        for row in _loop:
            mp_calc_stats((row, field, ds, stats, result_set), preprocess)
        # End for

        stat_results = pd.DataFrame.from_dict(result_set, orient='index')
        stat_results[field] = stat_results.index

        res_shp = shp.merge(stat_results, on=field)
        comment = "# Extracted from {} using {}".format(sheet_name,
                                                        shp_name)
        if preprocess is not None:
            comment += "\n# Preprocessed using {} : {}".format(*preprocess)

        results[(shp_fn, rst_fn)] = res_shp, sheet_name, comment
    # End for

    return results
# End extract_stats()

def apply_extract(d, shp, rst, field, stats, preprocess=None):
    from hazardstat import extract_stats
    d.update(extract_stats([shp], [rst], field, stats, preprocess))
    